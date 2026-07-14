from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from datetime import date, timedelta
from io import BytesIO
from app.database import attendance_collection, expenses_collection
from app.deps import get_current_user

router = APIRouter(prefix="/reports", tags=["reports"])


def _clean_attendance(docs):
    return [{
        "date": a["date"], "check_in": a.get("check_in"), "check_out": a.get("check_out"),
        "working_hours": a.get("working_hours", 0), "working_hours_display": a.get("working_hours_display", "0h 00m"),
        "overtime_hours": a.get("overtime_hours", 0), "overtime_display": a.get("overtime_display", "0h 00m"),
        "status": a.get("status", "Absent"), "work_note": a.get("work_note", ""),
    } for a in docs]


def _clean_expenses(docs):
    return [{
        "date": e["date"], "category": e["category"], "description": e["description"], "amount": e["amount"],
    } for e in docs]


async def _gather(current_user, start: date, end: date):
    attendance = await attendance_collection.find({
        "user_id": current_user["_id"],
        "date": {"$gte": start.isoformat(), "$lte": end.isoformat()},
    }).sort("date", 1).to_list(length=100)
    expenses = await expenses_collection.find({
        "user_id": current_user["_id"],
        "date": {"$gte": start.isoformat(), "$lte": end.isoformat()},
    }).sort("date", 1).to_list(length=1000)

    total_hours = round(sum(a.get("working_hours", 0) for a in attendance), 2)
    total_ot = round(sum(a.get("overtime_hours", 0) for a in attendance), 2)
    present_days = sum(1 for a in attendance if a.get("status") in ("Present", "Half Day"))
    total_expense = round(sum(e["amount"] for e in expenses), 2)
    overtime_amount = round(total_ot * current_user.get("overtime_rate", 0), 2)

    return {
        "attendance": _clean_attendance(attendance), "expenses": _clean_expenses(expenses),
        "total_hours": total_hours, "total_ot": total_ot,
        "present_days": present_days, "total_expense": total_expense,
        "overtime_amount": overtime_amount,
        "monthly_salary": current_user["monthly_salary"],
    }


@router.get("/weekly")
async def weekly_report(current_user: dict = Depends(get_current_user)):
    end = date.today()
    start = end - timedelta(days=6)
    return await _gather(current_user, start, end)


@router.get("/monthly")
async def monthly_report(current_user: dict = Depends(get_current_user)):
    end = date.today()
    start = end.replace(day=1)
    return await _gather(current_user, start, end)


@router.get("/export/pdf")
async def export_pdf(period: str = "monthly", current_user: dict = Depends(get_current_user)):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    end = date.today()
    start = end.replace(day=1) if period == "monthly" else end - timedelta(days=6)
    data = await _gather(current_user, start, end)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"{period.capitalize()} Report — {current_user['full_name']}", styles["Title"]),
        Spacer(1, 12),
        Paragraph(f"Period: {start} to {end}", styles["Normal"]),
        Spacer(1, 12),
        Paragraph(
            f"Total Working Hours: {data['total_hours']}h | Overtime: {data['total_ot']}h | "
            f"Present Days: {data['present_days']} | Total Expenses: ₹{data['total_expense']} | "
            f"Overtime Amount: ₹{data['overtime_amount']}",
            styles["Normal"],
        ),
        Spacer(1, 16),
        Paragraph("Attendance", styles["Heading2"]),
    ]

    att_rows = [["Date", "Check In", "Check Out", "Hours", "OT", "Status"]]
    for a in data["attendance"]:
        att_rows.append([a["date"], a.get("check_in") or "—", a.get("check_out") or "—",
                          a.get("working_hours_display", "0h 00m"), a.get("overtime_display", "0h 00m"),
                          a.get("status", "")])
    att_table = Table(att_rows, hAlign="LEFT")
    att_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(att_table)
    story.append(Spacer(1, 16))
    story.append(Paragraph("Expenses", styles["Heading2"]))

    exp_rows = [["Date", "Category", "Description", "Amount"]]
    for e in data["expenses"]:
        exp_rows.append([e["date"], e["category"], e["description"], f"₹{e['amount']}"])
    exp_table = Table(exp_rows, hAlign="LEFT")
    exp_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    story.append(exp_table)

    doc.build(story)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf", headers={
        "Content-Disposition": f"attachment; filename={period}_report.pdf"
    })


@router.get("/export/excel")
async def export_excel(period: str = "monthly", current_user: dict = Depends(get_current_user)):
    from openpyxl import Workbook

    end = date.today()
    start = end.replace(day=1) if period == "monthly" else end - timedelta(days=6)
    data = await _gather(current_user, start, end)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Attendance"
    ws1.append(["Date", "Check In", "Check Out", "Working Hours", "Overtime", "Status", "Work Note"])
    for a in data["attendance"]:
        ws1.append([a["date"], a.get("check_in") or "—", a.get("check_out") or "—",
                     a.get("working_hours_display", ""), a.get("overtime_display", ""),
                     a.get("status", ""), a.get("work_note", "")])

    ws2 = wb.create_sheet("Expenses")
    ws2.append(["Date", "Category", "Description", "Amount"])
    for e in data["expenses"]:
        ws2.append([e["date"], e["category"], e["description"], e["amount"]])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["Metric", "Value"])
    ws3.append(["Total Working Hours", data["total_hours"]])
    ws3.append(["Total Overtime Hours", data["total_ot"]])
    ws3.append(["Present Days", data["present_days"]])
    ws3.append(["Total Expenses", data["total_expense"]])
    ws3.append(["Overtime Amount", data["overtime_amount"]])
    ws3.append(["Monthly Salary", data["monthly_salary"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={period}_report.xlsx"},
    )
